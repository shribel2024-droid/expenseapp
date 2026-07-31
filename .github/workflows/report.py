"""
=========================================================
DailyExpense Manager
report.py

Report Generator

Compatible with:
- Python 3.11
- Pydroid 3
=========================================================
"""

import csv
import sqlite3
from datetime import datetime, timedelta

from database import DB_PATH

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph
    )
except ImportError:
    SimpleDocTemplate = None


class ReportManager:

    def __init__(self):
        self.db = DB_PATH

    # -------------------------------------------------
    # Database Connection
    # -------------------------------------------------

    def connect(self):
        return sqlite3.connect(self.db)

    # -------------------------------------------------
    # Execute Query
    # -------------------------------------------------

    def query(self, sql, values=()):

        conn = self.connect()
        cur = conn.cursor()

        cur.execute(sql, values)

        rows = cur.fetchall()

        conn.close()

        return rows

    # -------------------------------------------------
    # Total Expense
    # -------------------------------------------------

    def total_expense(self):

        data = self.query(
            "SELECT SUM(amount) FROM expenses"
        )

        if data[0][0] is None:
            return 0.0

        return float(data[0][0])

    # -------------------------------------------------
    # Total Transactions
    # -------------------------------------------------

    def total_transactions(self):

        data = self.query(
            "SELECT COUNT(*) FROM expenses"
        )

        return int(data[0][0])

    # -------------------------------------------------
    # Daily Report
    # -------------------------------------------------

    def daily_report(self, report_date=None):

        if report_date is None:
            report_date = datetime.now().strftime("%d-%m-%Y")

        sql = """
        SELECT
            expense_date,
            category,
            description,
            amount,
            payment_method
        FROM expenses
        WHERE expense_date=?
        ORDER BY id
        """

        return self.query(sql, (report_date,))

    # -------------------------------------------------
    # Weekly Report
    # -------------------------------------------------

    def weekly_report(self):

        today = datetime.today()

        start = today - timedelta(days=6)

        sql = """
        SELECT
            expense_date,
            category,
            description,
            amount
        FROM expenses
        """

        rows = self.query(sql)

        result = []

        for row in rows:

            try:

                d = datetime.strptime(
                    row[0],
                    "%d-%m-%Y"
                )

                if start.date() <= d.date() <= today.date():
                    result.append(row)

            except Exception:
                pass

        return result
        
            # -------------------------------------------------
    # Monthly Report
    # -------------------------------------------------

    def monthly_report(self, month=None, year=None):

        today = datetime.today()

        if month is None:
            month = today.month

        if year is None:
            year = today.year

        sql = """
        SELECT
            expense_date,
            category,
            description,
            amount,
            payment_method,
            notes
        FROM expenses
        """

        rows = self.query(sql)

        result = []

        for row in rows:

            try:
                d = datetime.strptime(row[0], "%d-%m-%Y")

                if d.month == month and d.year == year:
                    result.append(row)

            except Exception:
                continue

        return result

    # -------------------------------------------------
    # Monthly Expense Total
    # -------------------------------------------------

    def monthly_total(self, month=None, year=None):

        records = self.monthly_report(month, year)

        total = 0.0

        for row in records:
            total += float(row[3])

        return round(total, 2)

    # -------------------------------------------------
    # Monthly Transaction Count
    # -------------------------------------------------

    def monthly_transactions(self, month=None, year=None):

        return len(self.monthly_report(month, year))
            # -------------------------------------------------
    # Yearly Report
    # -------------------------------------------------

    def yearly_report(self, year=None):

        today = datetime.today()

        if year is None:
            year = today.year

        sql = """
        SELECT
            expense_date,
            category,
            description,
            amount,
            payment_method,
            notes
        FROM expenses
        """

        rows = self.query(sql)

        result = []

        for row in rows:

            try:
                d = datetime.strptime(row[0], "%d-%m-%Y")

                if d.year == year:
                    result.append(row)

            except Exception:
                continue

        return result

    # -------------------------------------------------
    # Yearly Expense Total
    # -------------------------------------------------

    def yearly_total(self, year=None):

        records = self.yearly_report(year)

        total = 0.0

        for row in records:
            total += float(row[3])

        return round(total, 2)

    # -------------------------------------------------
    # Yearly Transaction Count
    # -------------------------------------------------

    def yearly_transactions(self, year=None):

        return len(self.yearly_report(year))

    # -------------------------------------------------
    # Month-wise Summary
    # -------------------------------------------------

    def monthly_summary(self, year=None):

        today = datetime.today()

        if year is None:
            year = today.year

        summary = {}

        for month in range(1, 13):
            summary[month] = self.monthly_total(month, year)

        return summary
        
            # -------------------------------------------------
    # Category-wise Expense Summary
    # -------------------------------------------------

    def category_summary(self, start_date=None, end_date=None):
        """
        Returns category-wise expense totals.

        Parameters
        ----------
        start_date : str (DD-MM-YYYY) or None
        end_date   : str (DD-MM-YYYY) or None

        Returns
        -------
        [
            {
                "category": "...",
                "transactions": n,
                "total": amount
            },
            ...
        ]
        """

        rows = self.query("""
            SELECT
                expense_date,
                category,
                amount
            FROM expenses
        """)

        summary = {}

        for expense_date, category, amount in rows:

            try:
                dt = datetime.strptime(expense_date, "%d-%m-%Y")

                if start_date:
                    s = datetime.strptime(start_date, "%d-%m-%Y")
                    if dt < s:
                        continue

                if end_date:
                    e = datetime.strptime(end_date, "%d-%m-%Y")
                    if dt > e:
                        continue

            except Exception:
                continue

            if category not in summary:
                summary[category] = {
                    "transactions": 0,
                    "total": 0.0
                }

            summary[category]["transactions"] += 1
            summary[category]["total"] += float(amount)

        result = []

        for category in sorted(summary.keys()):
            result.append({
                "category": category,
                "transactions": summary[category]["transactions"],
                "total": round(summary[category]["total"], 2)
            })

        return result

    # -------------------------------------------------
    # Print Category Summary
    # -------------------------------------------------

    def print_category_summary(self, start_date=None, end_date=None):

        data = self.category_summary(start_date, end_date)

        print("=" * 70)
        print("{:<20} {:>15} {:>20}".format(
            "Category",
            "Transactions",
            "Total"
        ))
        print("=" * 70)

        grand_total = 0.0

        for item in data:

            print("{:<20} {:>15} {:>20.2f}".format(
                item["category"],
                item["transactions"],
                item["total"]
            ))

            grand_total += item["total"]

        print("=" * 70)
        print("{:<20} {:>15} {:>20.2f}".format(
            "Grand Total",
            "",
            grand_total
        ))
        
            # -------------------------------------------------
    # Payment Method Summary
    # -------------------------------------------------

    def payment_method_summary(self, start_date=None, end_date=None):
        """
        Returns payment method-wise expense summary.

        Parameters
        ----------
        start_date : DD-MM-YYYY (optional)
        end_date   : DD-MM-YYYY (optional)

        Returns
        -------
        [
            {
                "payment_method": "Cash",
                "transactions": 12,
                "total": 4250.00,
                "percentage": 35.40
            }
        ]
        """

        rows = self.query("""
            SELECT
                expense_date,
                payment_method,
                amount
            FROM expenses
        """)

        summary = {}
        grand_total = 0.0

        for expense_date, payment_method, amount in rows:

            try:
                dt = datetime.strptime(expense_date, "%d-%m-%Y")

                if start_date:
                    s = datetime.strptime(start_date, "%d-%m-%Y")
                    if dt < s:
                        continue

                if end_date:
                    e = datetime.strptime(end_date, "%d-%m-%Y")
                    if dt > e:
                        continue

            except Exception:
                continue

            if not payment_method:
                payment_method = "Unknown"

            value = float(amount)

            if payment_method not in summary:
                summary[payment_method] = {
                    "transactions": 0,
                    "total": 0.0
                }

            summary[payment_method]["transactions"] += 1
            summary[payment_method]["total"] += value

            grand_total += value

        result = []

        for method in sorted(summary.keys()):

            total = summary[method]["total"]

            percentage = 0

            if grand_total > 0:
                percentage = round((total / grand_total) * 100, 2)

            result.append({
                "payment_method": method,
                "transactions": summary[method]["transactions"],
                "total": round(total, 2),
                "percentage": percentage
            })

        return result


    # -------------------------------------------------
    # Print Payment Method Summary
    # -------------------------------------------------

    def print_payment_method_summary(
        self,
        start_date=None,
        end_date=None
    ):

        data = self.payment_method_summary(
            start_date,
            end_date
        )

        print("=" * 85)
        print("{:<20} {:>15} {:>18} {:>15}".format(
            "Payment Method",
            "Transactions",
            "Amount",
            "%"
        ))
        print("=" * 85)

        grand_total = 0

        for item in data:

            print("{:<20} {:>15} {:>18.2f} {:>14.2f}".format(
                item["payment_method"],
                item["transactions"],
                item["total"],
                item["percentage"]
            ))

            grand_total += item["total"]

        print("=" * 85)

        print("{:<20} {:>15} {:>18.2f}".format(
            "Grand Total",
            "",
            grand_total
        ))
        
            # -------------------------------------------------
    # Budget Utilization
    # -------------------------------------------------

    def budget_utilization(self, month=None, year=None):
        """
        Calculate budget utilization for a given month.

        Returns a dictionary containing:
        - budget
        - spent
        - remaining
        - percentage_used
        - status
        """

        today = datetime.today()

        if month is None:
            month = today.month

        if year is None:
            year = today.year

        # Get Budget
        budget_row = self.query("""
            SELECT budget_amount
            FROM budget
            WHERE month=? AND year=?
            LIMIT 1
        """, (month, year))

        if budget_row:
            budget = float(budget_row[0][0])
        else:
            budget = 0.0

        # Monthly Expense
        spent = self.monthly_total(month, year)

        remaining = budget - spent

        if budget > 0:
            percentage = round((spent / budget) * 100, 2)
        else:
            percentage = 0.0

        # Budget Status
        if budget == 0:
            status = "Budget Not Set"
        elif spent > budget:
            status = "Over Budget"
        elif percentage >= 90:
            status = "Critical"
        elif percentage >= 75:
            status = "Warning"
        else:
            status = "Within Budget"

        return {
            "month": month,
            "year": year,
            "budget": round(budget, 2),
            "spent": round(spent, 2),
            "remaining": round(remaining, 2),
            "percentage_used": percentage,
            "status": status
        }

    # -------------------------------------------------
    # Print Budget Utilization
    # -------------------------------------------------

    def print_budget_utilization(self, month=None, year=None):

        data = self.budget_utilization(month, year)

        print("=" * 50)
        print("BUDGET UTILIZATION")
        print("=" * 50)
        print(f"Month            : {data['month']:02d}")
        print(f"Year             : {data['year']}")
        print(f"Budget           : ₹ {data['budget']:.2f}")
        print(f"Spent            : ₹ {data['spent']:.2f}")
        print(f"Remaining        : ₹ {data['remaining']:.2f}")
        print(f"Used             : {data['percentage_used']:.2f}%")
        print(f"Status           : {data['status']}")
        print("=" * 50)

    # -------------------------------------------------
    # Set / Update Monthly Budget
    # -------------------------------------------------

    def set_budget(self, month, year, amount):

        existing = self.query("""
            SELECT id
            FROM budget
            WHERE month=? AND year=?
        """, (month, year))

        conn = self.connect()
        cur = conn.cursor()

        if existing:
            cur.execute("""
                UPDATE budget
                SET budget_amount=?
                WHERE month=? AND year=?
            """, (amount, month, year))
        else:
            cur.execute("""
                INSERT INTO budget
                (month, year, budget_amount)
                VALUES (?, ?, ?)
            """, (month, year, amount))

        conn.commit()
        conn.close()
        
            # -------------------------------------------------
    # Export Report to CSV
    # -------------------------------------------------

    def export_csv(
        self,
        filename="expense_report.csv",
        report_type="all",
        month=None,
        year=None
    ):
        """
        Export expense data to CSV.

        report_type:
            all
            daily
            weekly
            monthly
            yearly
        """

        report_type = report_type.lower()

        if report_type == "daily":
            records = self.daily_report()

        elif report_type == "weekly":
            records = self.weekly_report()

        elif report_type == "monthly":
            records = self.monthly_report(month, year)

        elif report_type == "yearly":
            records = self.yearly_report(year)

        else:
            records = self.query("""
                SELECT
                    expense_date,
                    expense_time,
                    category,
                    description,
                    amount,
                    payment_method,
                    notes
                FROM expenses
                ORDER BY expense_date
            """)

        try:

            with open(
                filename,
                "w",
                newline="",
                encoding="utf-8"
            ) as csvfile:

                writer = csv.writer(csvfile)

                writer.writerow([
                    "Date",
                    "Time",
                    "Category",
                    "Description",
                    "Amount",
                    "Payment Method",
                    "Notes"
                ])

                total = 0.0

                for row in records:

                    writer.writerow(row)

                    try:
                        # Amount column differs in summary reports
                        if len(row) >= 7:
                            total += float(row[4])
                        else:
                            total += float(row[3])
                    except Exception:
                        pass

                writer.writerow([])
                writer.writerow(["Total Expense", "", "", "", total])

            return True, filename

        except Exception as e:
            return False, str(e)

    # -------------------------------------------------
    # Export Category Summary to CSV
    # -------------------------------------------------

    def export_category_summary_csv(
        self,
        filename="category_summary.csv",
        start_date=None,
        end_date=None
    ):

        data = self.category_summary(start_date, end_date)

        try:

            with open(
                filename,
                "w",
                newline="",
                encoding="utf-8"
            ) as file:

                writer = csv.writer(file)

                writer.writerow([
                    "Category",
                    "Transactions",
                    "Total"
                ])

                grand_total = 0.0

                for item in data:

                    writer.writerow([
                        item["category"],
                        item["transactions"],
                        item["total"]
                    ])

                    grand_total += item["total"]

                writer.writerow([])
                writer.writerow([
                    "Grand Total",
                    "",
                    grand_total
                ])

            return True, filename

        except Exception as e:
            return False, str(e)
            
                # -------------------------------------------------
    # Export Report to Excel
    # -------------------------------------------------

    def export_excel(
        self,
        filename="expense_report.xlsx",
        report_type="all",
        month=None,
        year=None
    ):
        """
        Export expense report to Microsoft Excel (.xlsx)

        report_type:
            all
            daily
            weekly
            monthly
            yearly
        """

        if Workbook is None:
            return False, "openpyxl library is not installed."

        report_type = report_type.lower()

        if report_type == "daily":
            records = self.daily_report()

        elif report_type == "weekly":
            records = self.weekly_report()

        elif report_type == "monthly":
            records = self.monthly_report(month, year)

        elif report_type == "yearly":
            records = self.yearly_report(year)

        else:
            records = self.query("""
                SELECT
                    expense_date,
                    expense_time,
                    category,
                    description,
                    amount,
                    payment_method,
                    notes
                FROM expenses
                ORDER BY expense_date
            """)

        try:

            wb = Workbook()
            ws = wb.active
            ws.title = "Expense Report"

            headers = [
                "Date",
                "Time",
                "Category",
                "Description",
                "Amount",
                "Payment Method",
                "Notes"
            ]

            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header

            total = 0.0

            for row_num, row in enumerate(records, start=2):

                for col_num, value in enumerate(row, start=1):
                    ws.cell(row=row_num, column=col_num).value = value

                try:
                    if len(row) >= 7:
                        total += float(row[4])
                    else:
                        total += float(row[3])
                except Exception:
                    pass

            summary_row = len(records) + 3

            ws.cell(summary_row, 1).value = "Total Expense"
            ws.cell(summary_row, 5).value = total

            wb.save(filename)

            return True, filename

        except Exception as e:
            return False, str(e)


    # -------------------------------------------------
    # Export Category Summary to Excel
    # -------------------------------------------------

    def export_category_summary_excel(
        self,
        filename="category_summary.xlsx",
        start_date=None,
        end_date=None
    ):

        if Workbook is None:
            return False, "openpyxl library is not installed."

        data = self.category_summary(start_date, end_date)

        try:

            wb = Workbook()
            ws = wb.active
            ws.title = "Category Summary"

            ws.append([
                "Category",
                "Transactions",
                "Total Amount"
            ])

            grand_total = 0.0

            for item in data:

                ws.append([
                    item["category"],
                    item["transactions"],
                    item["total"]
                ])

                grand_total += item["total"]

            ws.append([])
            ws.append([
                "Grand Total",
                "",
                grand_total
            ])

            wb.save(filename)

            return True, filename

        except Exception as e:
            return False, str(e)


    # -------------------------------------------------
    # Export Payment Summary to Excel
    # -------------------------------------------------

    def export_payment_summary_excel(
        self,
        filename="payment_summary.xlsx",
        start_date=None,
        end_date=None
    ):

        if Workbook is None:
            return False, "openpyxl library is not installed."

        data = self.payment_method_summary(
            start_date,
            end_date
        )

        try:

            wb = Workbook()
            ws = wb.active
            ws.title = "Payment Summary"

            ws.append([
                "Payment Method",
                "Transactions",
                "Amount",
                "Percentage"
            ])

            grand_total = 0.0

            for item in data:

                ws.append([
                    item["payment_method"],
                    item["transactions"],
                    item["total"],
                    item["percentage"]
                ])

                grand_total += item["total"]

            ws.append([])
            ws.append([
                "Grand Total",
                "",
                grand_total,
                ""
            ])

            wb.save(filename)

            return True, filename

        except Exception as e:
            return False, str(e)
            
                # -------------------------------------------------
    # Export PDF Report
    # -------------------------------------------------

    def export_pdf(
        self,
        filename="expense_report.pdf",
        report_type="all",
        month=None,
        year=None
    ):

        if SimpleDocTemplate is None:
            return False, "ReportLab library is not installed."

        report_type = report_type.lower()

        if report_type == "daily":
            records = self.daily_report()

        elif report_type == "weekly":
            records = self.weekly_report()

        elif report_type == "monthly":
            records = self.monthly_report(month, year)

        elif report_type == "yearly":
            records = self.yearly_report(year)

        else:
            records = self.query("""
                SELECT
                    expense_date,
                    expense_time,
                    category,
                    description,
                    amount,
                    payment_method,
                    notes
                FROM expenses
                ORDER BY expense_date
            """)

        try:

            doc = SimpleDocTemplate(filename)

            styles = getSampleStyleSheet()

            story = []

            story.append(
                Paragraph(
                    "<b>DailyExpense Manager</b>",
                    styles["Title"]
                )
            )

            story.append(
                Paragraph(
                    f"Report Type : {report_type.title()}",
                    styles["Normal"]
                )
            )

            story.append(
                Paragraph(
                    f"Generated : {datetime.now().strftime('%d-%m-%Y %H:%M')}",
                    styles["Normal"]
                )
            )

            story.append(
                Paragraph("<br/>", styles["Normal"])
            )

            table_data = [[
                "Date",
                "Time",
                "Category",
                "Description",
                "Amount",
                "Payment"
            ]]

            total = 0.0

            for row in records:

                if len(row) >= 7:

                    table_data.append([
                        row[0],
                        row[1],
                        row[2],
                        row[3],
                        f"{float(row[4]):.2f}",
                        row[5]
                    ])

                    total += float(row[4])

                else:

                    table_data.append([
                        row[0],
                        "",
                        row[1],
                        row[2],
                        f"{float(row[3]):.2f}",
                        ""
                    ])

                    total += float(row[3])

            table = Table(table_data)

            table.setStyle(TableStyle([

                ('BACKGROUND', (0,0), (-1,0), colors.grey),

                ('TEXTCOLOR', (0,0), (-1,0), colors.white),

                ('ALIGN', (0,0), (-1,-1), 'CENTER'),

                ('GRID', (0,0), (-1,-1), 0.5, colors.black),

                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),

                ('BOTTOMPADDING', (0,0), (-1,0), 8),

                ('BACKGROUND', (0,1), (-1,-1), colors.beige)

            ]))

            story.append(table)

            story.append(
                Paragraph("<br/>", styles["Normal"])
            )

            story.append(
                Paragraph(
                    f"<b>Total Expense : ₹ {total:.2f}</b>",
                    styles["Heading2"]
                )
            )

            story.append(
                Paragraph(
                    f"<b>Total Transactions : {len(records)}</b>",
                    styles["Heading2"]
                )
            )

            doc.build(story)

            return True, filename

        except Exception as e:

            return False, str(e)